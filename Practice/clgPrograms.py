# 1a
m1 = int(input("Enter the marks in the first test: "))
m2 = int(input("Enter the marks in second test: "))
m3 = int(input("Enter the marks in third test: "))
if m1 > m2:
    if m2 > m3:
        total = m1 + m2
    else:
        total = m1 + m3
elif m1 > m3:
    total = m1 + m2
else:
    total = m2 + m3
Avg = total / 2
print("The average of the best two test marks is: ", Avg)


# 1b
val = int(input("Enter a value : "))
str_val = str(val)
if str_val == str_val[::-1]:
    print("Palindrome")
    for i in range(10):
        if str_val.count(str(i)) > 0:
            print(str(i), "appears", str_val.count(str(i)), "times")
else:
    print("Not Palindrome")


# 2a
def fn(n):
    if n == 1:
        return 0
    elif n == 2:
        return 1
    else:
        return fn(n - 1) + fn(n - 2)


num = int(input("Enter a number : "))
if num > 0:
    print("fn(", num, ") = ", fn(num), sep="")
else:
    print("Error in input")


# 2b
def bin2Dec(val):
    rev = val[::-1]
    dec = 0
    i = 0
    for dig in rev:
        dec += int(dig) * 2**i
        i += 1
    return dec


def oct2Hex(val):
    rev = val[::-1]
    dec = 0
    i = 0
    for dig in rev:
        dec += int(dig) * 8**i
        i += 1
    list = []
    while dec != 0:
        list.append(dec % 16)
        dec = dec // 16
    nl = []
    for elem in list[::-1]:
        if elem <= 9:
            nl.append(str(elem))
        else:
            nl.append(chr(ord("A") + (elem - 10)))
    hex = "".join(nl)
    return hex


num1 = input("Enter a binary number : ")
print(bin2Dec(num1))
num2 = input("Enter a octal number : ")
print(oct2Hex(num2))


# 3a
sentence = input("Enter a sentence : ")
wordList = sentence.split(" ")
print("This sentence has", len(wordList), "words")
digCnt = upCnt = loCnt = 0
for ch in sentence:
    if "0" <= ch <= "9":
        digCnt += 1
    elif "A" <= ch <= "Z":
        upCnt += 1
    elif "a" <= ch <= "z":
        loCnt += 1
print(
    "This sentence has",
    digCnt,
    "digits",
    upCnt,
    "upper case letters",
    loCnt,
    "lower case letters",
)


# 3b
str1 = input("Enter String 1 \n")
str2 = input("Enter String 2 \n")
if len(str2) < len(str1):
    short = len(str2)
    long = len(str1)
else:
    short = len(str1)
    long = len(str2)
matchCnt = 0
for i in range(short):
    if str1[i] == str2[i]:
        matchCnt += 1
print("Similarity between two said strings:")
print(matchCnt / long)


# 4a
import random


def merge_sort(lst):
    if len(lst) > 1:
        mid = len(lst) // 2
        left_half = lst[:mid]
        right_half = lst[mid:]
        merge_sort(left_half)
        merge_sort(right_half)
        i = j = k = 0
        while i < len(left_half) and j < len(right_half):
            if left_half[i] < right_half[j]:
                lst[k] = left_half[i]
                i += 1
            else:
                lst[k] = right_half[j]
                j += 1
            k += 1
        while i < len(left_half):
            lst[k] = left_half[i]
            i += 1
            k += 1
        while j < len(right_half):
            lst[k] = right_half[j]
            j += 1
            k += 1
    return lst


def insertion_sort(arr):
    for i in range(1, len(arr)):
        key = arr[i]
        j = i - 1
        while j >= 0 and key < arr[j]:
            arr[j + 1] = arr[j]
            j -= 1
        arr[j + 1] = key


my_list = []
for i in range(10):
    my_list.append(random.randint(0, 999))
print("\nUnsorted List")
print(my_list)
print("Sorting using Insertion Sort")
insertion_sort(my_list)
print(my_list)
my_list = []
for i in range(10):
    my_list.append(random.randint(0, 999))
print("\nUnsorted List")
print(my_list)
print("Sorting using Merge Sort")
merge_sort(my_list)
print(my_list)


# 4b
def roman2Dec(romStr):
    roman_dict = {"I": 1, "V": 5, "X": 10, "L": 50, "C": 100, "D": 500, "M": 1000}
    # Analyze string backwards
    romanBack = list(romStr)[::-1]
    value = 0
    # To keep track of order
    rightVal = roman_dict[romanBack[0]]
    for numeral in romanBack:
        leftVal = roman_dict[numeral]
        # Check for subtraction
        if leftVal < rightVal:
            value -= leftVal
        else:
            value += leftVal
        rightVal = leftVal
    return value


romanStr = input("Enter a Roman Number : ")
print(roman2Dec(romanStr))


# 5a
import re


def isphonenumber(numStr):
    if len(numStr) != 12:
        return False
    for i in range(len(numStr)):
        if i == 3 or i == 7:
            if numStr[i] != "-":
                return False
            else:
                if numStr[i].isdigit() == False:
                    return False
    return True


def chkphonenumber(numStr):
    ph_no_pattern = re.compile(r"^\d{3}-\d{3}-\d{4}$")
    if ph_no_pattern.match(numStr):
        return True
    else:
        return False


ph_num = input("Enter a phone number : ")
print("Without using Regular Expression")
if isphonenumber(ph_num):
    print("Valid phone number")
else:
    print("Invalid phone number")
print("Using Regular Expression")
if chkphonenumber(ph_num):
    print("Valid phone number")
else:
    print("Invalid phone number")


# 5b
import re

phone_regex = re.compile(r"\+\d{12}")
email_regex = re.compile(r"[A-Za-z0-9._]+@[A-Za-z0-9]+\.[A-Z|a-z]{2,}")
# Open the file for reading
with open("example.txt", "r") as f:
    # Loop through each line in the file
    for line in f:
        # Search for phone numbers in the line
        matches = phone_regex.findall(line)
        # Print any matches found
        for match in matches:
            print(match)
        matches = email_regex.findall(line)
        # Print any matches found
        for match in matches:
            print(match)


# 6a
import os.path
import sys

fname = input("Enter the filename : ")
if not os.path.isfile(fname):
    print("File", fname, "doesn't exists")
    sys.exit(0)
infile = open(fname, "r")
lineList = infile.readlines()
for i in range(20):
    print(i + 1, ":", lineList[i])
word = input("Enter a word : ")
cnt = 0
for line in lineList:
    cnt += line.count(word)
print("The word", word, "appears", cnt, "times in the file")


# 6b
import shutil

name = input("Enter zip file name: ")
directory = input("Enter file directory: ")
shutil.make_archive(name, "zip", directory)


# 9b
from openpyxl import Workbook
from openpyxl.styles import Font

wb = Workbook()
sheet = wb.active
sheet.title = "Language"
wb.create_sheet(title="Capital")
lang = ["Kannada", "Telugu", "Tamil"]
state = ["Karnataka", "Telangana", "Tamil Nadu"]
capital = ["Bengaluru", "Hyderabad", "Chennai"]
code = ["KA", "TS", "TN"]
sheet.cell(row=1, column=1).value = "State"
sheet.cell(row=1, column=2).value = "Language"
sheet.cell(row=1, column=3).value = "Code"
sheet.cell(row=1, column=4).value = "Capital"
ft = Font(bold=True)
for row in sheet["A1:D1"]:
    for cell in row:
        cell.font = ft
for i in range(2, 5):
    sheet.cell(row=i, column=1).value = state[i - 2]
    sheet.cell(row=i, column=2).value = lang[i - 2]
    sheet.cell(row=i, column=3).value = code[i - 2]
    sheet.cell(row=i, column=4).value = capital[i - 2]
wb.save("demo.xlsx")
srchCode = input("Enter state code for finding capital ")
for i in range(2, 5):
    data = sheet.cell(row=i, column=3).value
    if data == srchCode:
        print(
            "Corresponding capital for code",
            srchCode,
            "is",
            sheet.cell(row=i, column=4).value,
        )
srchCode = input("Enter state code for finding language ")
for i in range(2, 5):
    data = sheet.cell(row=i, column=3).value
    if data == srchCode:
        print(
            "Corresponding language for code",
            srchCode,
            "is",
            sheet.cell(row=i, column=2).value,
        )
wb.close()


# 10a
from PyPDF2 import PdfWriter, PdfReader

num = int(input("Enter page number you want combine from multiple documents "))
pdf1 = open("birds.pdf", "rb")
pdf2 = open("birdspic.pdf", "rb")
pdf_writer = PdfWriter()
pdf1_reader = PdfReader(pdf1)
page = pdf1_reader.pages[num - 1]
pdf_writer.add_page(page)
pdf2_reader = PdfReader(pdf2)
page = pdf2_reader.pages[num - 1]
pdf_writer.add_page(page)
with open("output.pdf", "wb") as output:
    pdf_writer.write(output)


# 10b
import json

# Load the JSON data from file
with open("weather_data.json") as f:
    data = json.load(f)
# Extract the required weather data
current_temp = data["main"]["temp"]
humidity = data["main"]["humidity"]
weather_desc = data["weather"][0]["description"]
# Display the weather data
print(f"Current temperature: {current_temp}°C")
print(f"Humidity: {humidity}%")
print(f"Weather description: {weather_desc}")


# 9a
import requests
import os
from bs4 import BeautifulSoup

# Set the URL of the first XKCD comic
url = "https://xkcd.com/1/"

# Create a folder to store the comics
if not os.path.exists("xkcd_comics"):
    os.makedirs("xkcd_comics")

# Loop through all the comics
while True:
    # Download the page content
    res = requests.get(url)
    res.raise_for_status()

    # Parse the page content using BeautifulSoup
    soup = BeautifulSoup(res.text, "html.parser")

    # Find the URL of the comic image
    comic_elem = soup.select("#comic img")
    if comic_elem == []:
        print("Could not find comic image.")
    else:
        comic_url = "https:" + comic_elem[0].get("src")

        # Download the comic image
        print(f"Downloading {comic_url}...")
        res = requests.get(comic_url)
        res.raise_for_status()

        # Save the comic image to the xkcd_comics folder
        image_file = open(
            os.path.join("xkcd_comics", os.path.basename(comic_url)), "wb"
        )
        for chunk in res.iter_content(100000):
            image_file.write(chunk)
        image_file.close()

    # Get the URL of the previous comic
    prev_link = soup.select('a[rel="prev"]')[0]
    if prev_link is None:
        break
    url = "https://xkcd.com" + prev_link.get("href")

print("All comics downloaded.")


# 7a
import math


class Shape:
    def __init__(self):
        self.area = 0
        self.name = ""

    def showArea(self):
        print("The area of the", self.name, "is", self.area, "units")


class Circle(Shape):
    def __init__(self, radius):
        super().__init__()
        self.area = 0
        self.name = "Circle"
        self.radius = radius

    def calcArea(self):
        self.area = math.pi * self.radius * self.radius


class Rectangle(Shape):
    def __init__(self, length, breadth):
        super().__init__()
        self.area = 0
        self.name = "Rectangle"
        self.length = length
        self.breadth = breadth

    def calcArea(self):
        self.area = self.length * self.breadth


class Triangle(Shape):
    def __init__(self, base, height):
        super().__init__()
        self.area = 0
        self.name = "Triangle"
        self.base = base
        self.height = height

    def calcArea(self):
        self.area = self.base * self.height / 2


c1 = Circle(5)
c1.calcArea()
c1.showArea()
r1 = Rectangle(5, 4)
r1.calcArea()
r1.showArea()
t1 = Triangle(3, 4)
t1.calcArea()
t1.showArea()


# 7b
class Employee:
    def __init__(self):
        self.name = ""
        self.empId = ""
        self.dept = ""
        self.salary = 0

    def getEmpDetails(self):
        self.name = input("Enter Employee name : ")
        self.empId = input("Enter Employee ID : ")
        self.dept = input("Enter Employee Dept : ")
        self.salary = int(input("Enter Employee Salary : "))

    def showEmpDetails(self):
        print("Employee Details")
        print("Name : ", self.name)
        print("ID : ", self.empId)
        print("Dept : ", self.dept)
        print("Salary : ", self.salary)

    def updtSalary(self):
        self.salary = int(input("Enter new Salary : "))
        print("Updated Salary", self.salary)


e1 = Employee()
e1.getEmpDetails()
e1.showEmpDetails()
e1.updtSalary()


# 8
class PaliStr:
    def __init__(self):
        self.isPali = False

    def chkPalindrome(self, myStr):
        if myStr == myStr[::-1]:
            self.isPali = True
        else:
            self.isPali = False
        return self.isPali


class PaliInt(PaliStr):
    def __init__(self):
        super().__init__()
        self.isPali = False

    def chkPalindrome(self, val):
        temp = val
        rev = 0
        while temp != 0:
            dig = temp % 10
            rev = (rev * 10) + dig
            temp = temp // 10
        if val == rev:
            self.isPali = True
        else:
            self.isPali = False
        return self.isPali


st = input("Enter a string : ")
stObj = PaliStr()
if stObj.chkPalindrome(st):
    print("Given string is a Palindrome")
else:
    print("Given string is not a Palindrome")
val = int(input("Enter a integer : "))
intObj = PaliInt()
if intObj.chkPalindrome(val):
    print("Given integer is a Palindrome")
else:
    print("Given integer is not a Palindrome")


# 10 A
from PyPDF2 import PdfWriter, PdfReader

num = int(input("Enter page number you want combine from multiple documents "))
pdf1 = open("birds.pdf", "rb")
pdf2 = open("birdspic.pdf", "rb")
pdf_writer = PdfWriter()
pdf1_reader = PdfReader(pdf1)
page = pdf1_reader.pages[num - 1]
pdf_writer.add_page(page)
pdf2_reader = PdfReader(pdf2)
page = pdf2_reader.pages[num - 1]
pdf_writer.add_page(page)
with open("output.pdf", "wb") as output:
    pdf_writer.write(output)
